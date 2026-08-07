import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import itertools
import pandas as pd
import urllib.request
import json
import sqlite3

st.set_page_config(page_title="Shopee Auto Price & Mass Upload Builder", layout="wide")

# --- พจนานุกรมรองรับ 2 ภาษา (TH / EN) ---
TEXTS = {
    "TH": {
        "title": "📦 เครื่องมือสร้างไฟล์ Mass Upload Shopee & คำนวณราคาขายอัตโนมัติ (มี Auto-save 💾)",
        "calc_setting": "⚙️ ตั้งค่าการคำนวณราคา (Target Market & Real-time Exchange Rate)",
        "currency_select": "เลือกตลาดเป้าหมาย",
        "rate_label": "อัตราแลกเปลี่ยน Real-time (1 {curr} -> JPY)",
        "rate_info": "💡 ดึงข้อมูลอัตราแลกเปลี่ยน Real-time ล่าสุดอัตโนมัติ",
        "add_product": "➕ เพิ่มสินค้าชิ้นใหม่",
        "clear_all": "🗑️ ล้างข้อมูลสินค้าทั้งหมด",
        "del_product": "🗑️ ลบสินค้านี้",
        "product_num": "🛒 สินค้าชิ้นที่",
        "cat_id": "Category ID / รหัสหมวดหมู่",
        "parent_sku": "Parent SKU / รหัสอ้างอิงหลัก",
        "integration_no": "Variation Integration No.",
        "brand": "แบรนด์ (Brand)",
        "p_name": "ชื่อสินค้า",
        "weight": "น้ำหนักสินค้าเริ่มต้น (g)",
        "p_desc": "รายละเอียดสินค้า",
        "cover_img": "URL รูปภาพปกหลัก & รูปสินค้า (คั่นด้วย ,)",
        "cover_img_help": "ใส่ URL คั่นด้วยเครื่องหมายจุลภาค รูปแรก = Cover Image",
        "v1_name": "ชื่อตัวเลือกที่ 1 (เช่น สี / รุ่น)",
        "v1_opts": "รายการตัวเลือกที่ 1 (คั่นด้วย ,)",
        "v1_imgs": "URL รูปภาพตัวเลือกที่ 1 (คั่นด้วย ,)",
        "v1_imgs_help": "ระบุ URL รูปตามลำดับตัวเลือกที่ 1",
        "v2_name": "ชื่อตัวเลือกที่ 2 (เช่น ไซส์) [เว้นว่างได้]",
        "v2_opts": "รายการตัวเลือกที่ 2 (คั่นด้วย ,)",
        "batch_title": "⚡ ตั้งค่าด่วน (แยกปรับแต่ละค่าไปยังทุก Variation):",
        "btn_apply": "⚡ นำไปใช้",
        "grid_title": "💰 ตารางกำหนดราคาซื้อ (JPY), น้ำหนัก (g), Profit Rate (%), สต๊อก และราคาขาย:",
        "cost_col": "Buying Price (JPY)",
        "weight_col": "Weight (g)",
        "profit_col": "Profit Rate (%)",
        "price_col": "Selling Price",
        "stock_col": "Stock (ชิ้น)",
        "sku_col": "SKU (แก้ไขได้)",
        "btn_generate": "🚀 สร้างไฟล์ Excel รวมทุกสินค้าสำหรับ Shopee",
        "success_msg": "✅ สร้างไฟล์สำเร็จ! รวมสินค้าทั้งหมด {count} รายการ",
        "btn_download": "📥 ดาวน์โหลดไฟล์ Excel พร้อมอัปโหลด Shopee",
        "default_pname": "รองเท้าสปอร์ตผ้าใบคุณภาพสูง",
        "default_pdesc": "รองเท้าสปอร์ต นุ่ม สวมใส่สบาย",
        "default_v1_name": "สี",
        "default_v1_opts": "WINE, WHITE",
        "default_v2_name": "ไซส์",
        "default_v2_opts": "23.0cm., 24.0cm., 25.0cm., 26.0cm.",
        "lang_select": "🌐 ภาษา / Language"
    },
    "EN": {
        "title": "📦 Shopee Mass Upload Builder & Auto Price Calculator (with Auto-save 💾)",
        "calc_setting": "⚙️ Price Calculation Settings (Target Market & Real-time Exchange Rate)",
        "currency_select": "Select Target Market",
        "rate_label": "Real-time Exchange Rate (1 {curr} -> JPY)",
        "rate_info": "💡 Auto-fetching latest real-time exchange rates",
        "add_product": "➕ Add New Product",
        "clear_all": "🗑️ Clear All Products",
        "del_product": "🗑️ Delete Product",
        "product_num": "🛒 Product #",
        "cat_id": "Category ID",
        "parent_sku": "Parent SKU",
        "integration_no": "Variation Integration No.",
        "brand": "Brand",
        "p_name": "Product Name",
        "weight": "Default Product Weight (g)",
        "p_desc": "Product Description",
        "cover_img": "Cover Image & Product URLs (comma separated)",
        "cover_img_help": "Separate URLs with commas. First image = Cover Image",
        "v1_name": "Variation 1 Name (e.g. Color)",
        "v1_opts": "Variation 1 Options (comma separated)",
        "v1_imgs": "Variation 1 Image URLs (comma separated)",
        "v1_imgs_help": "Specify image URLs matching Variation 1 order",
        "v2_name": "Variation 2 Name (e.g. Size) [Optional]",
        "v2_opts": "Variation 2 Options (comma separated)",
        "batch_title": "⚡ Quick Batch Apply (Apply values to all Variations):",
        "btn_apply": "⚡ Apply",
        "grid_title": "💰 Pricing, Weight (g), Profit Rate (%), Stock, and Selling Price Table:",
        "cost_col": "Buying Price (JPY)",
        "weight_col": "Weight (g)",
        "profit_col": "Profit Rate (%)",
        "price_col": "Selling Price",
        "stock_col": "Stock (pcs)",
        "sku_col": "SKU (Editable)",
        "btn_generate": "🚀 Generate Combined Excel File for Shopee",
        "success_msg": "✅ Generated successfully! Total products: {count}",
        "btn_download": "📥 Download Shopee Mass Upload Excel",
        "default_pname": "High Quality Sneakers",
        "default_pdesc": "Comfortable sports shoes",
        "default_v1_name": "Color",
        "default_v1_opts": "WINE, WHITE",
        "default_v2_name": "Size",
        "default_v2_opts": "23.0cm., 24.0cm., 25.0cm., 26.0cm.",
        "lang_select": "🌐 Language / ภาษา"
    }
}

# --- SIDEBAR LANGUAGE SELECTION ---
lang_choice = st.sidebar.selectbox("🌐 Language / ภาษา", ["TH", "EN"], key="app_language")
T = TEXTS[lang_choice]

st.title(T["title"])

# --- DATABASE SETUP ---
DB_FILE = "shopee_products.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS products (
            p_id INTEGER PRIMARY KEY, cat_id TEXT, psku TEXT, integ TEXT,
            brand TEXT, name TEXT, weight REAL, desc TEXT, cimg TEXT,
            v1n TEXT, v1o TEXT, v1i TEXT, v2n TEXT, v2o TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS variations (
            p_id INTEGER, variation_title TEXT, sku TEXT, cost_jpy INTEGER,
            weight_g REAL, profit_rate REAL, stock INTEGER, opt1 TEXT, opt2 TEXT,
            PRIMARY KEY (p_id, variation_title)
        )
    ''')
    conn.commit()
    conn.close()

def save_product_to_db(p_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO products 
        (p_id, cat_id, psku, integ, brand, name, weight, desc, cimg, v1n, v1o, v1i, v2n, v2o)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        p_id, st.session_state.get(f"cat_{p_id}", ""), st.session_state.get(f"psku_{p_id}", ""),
        st.session_state.get(f"integ_{p_id}", ""), st.session_state.get(f"brand_{p_id}", ""),
        st.session_state.get(f"name_{p_id}", ""), float(st.session_state.get(f"w_{p_id}", 0.0)),
        st.session_state.get(f"desc_{p_id}", ""), st.session_state.get(f"cimg_{p_id}", ""),
        st.session_state.get(f"v1n_{p_id}", ""), st.session_state.get(f"v1o_{p_id}", ""),
        st.session_state.get(f"v1i_{p_id}", ""), st.session_state.get(f"v2n_{p_id}", ""),
        st.session_state.get(f"v2o_{p_id}", "")
    ))
    conn.commit()
    conn.close()

def save_variations_to_db(p_id, df_vars):
    if df_vars is None: return
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM variations WHERE p_id = ?', (p_id,))
    if not df_vars.empty:
        for _, row in df_vars.iterrows():
            c.execute('''
                INSERT INTO variations (p_id, variation_title, sku, cost_jpy, weight_g, profit_rate, stock, opt1, opt2)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                p_id, str(row.get("Variation", "")), str(row.get("SKU", "")),
                int(row.get("cost_jpy", 0)), float(row.get("weight_g", 0)),
                float(row.get("profit_rate", 0)), int(row.get("stock", 0)),
                str(row.get("Opt1", "")), str(row.get("Opt2", ""))
            ))
    conn.commit()
    conn.close()

def load_all_from_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT * FROM products ORDER BY p_id ASC')
    prods = c.fetchall()
    if not prods:
        conn.close()
        return False

    st.session_state.products_list = []
    max_id = -1
    for row in prods:
        p_id = row[0]
        st.session_state.products_list.append(p_id)
        if p_id > max_id: max_id = p_id
            
        st.session_state[f"cat_{p_id}"] = row[1]
        st.session_state[f"psku_{p_id}"] = row[2]
        st.session_state[f"integ_{p_id}"] = row[3]
        st.session_state[f"brand_{p_id}"] = row[4]
        st.session_state[f"name_{p_id}"] = row[5]
        st.session_state[f"w_{p_id}"] = row[6]
        st.session_state[f"desc_{p_id}"] = row[7]
        st.session_state[f"cimg_{p_id}"] = row[8]
        st.session_state[f"v1n_{p_id}"] = row[9]
        st.session_state[f"v1o_{p_id}"] = row[10]
        st.session_state[f"v1i_{p_id}"] = row[11]
        st.session_state[f"v2n_{p_id}"] = row[12]
        st.session_state[f"v2o_{p_id}"] = row[13]

        c.execute('SELECT variation_title, sku, cost_jpy, weight_g, profit_rate, stock, opt1, opt2 FROM variations WHERE p_id = ?', (p_id,))
        vars_rows = c.fetchall()
        if vars_rows:
            var_list = []
            for v in vars_rows:
                var_list.append({
                    "Variation": v[0], "SKU": v[1], "cost_jpy": v[2],
                    "weight_g": v[3], "profit_rate": v[4], "selling_price": 0,
                    "stock": v[5], "Opt1": v[6], "Opt2": v[7]
                })
            st.session_state[f"df_data_{p_id}"] = pd.DataFrame(var_list)

    st.session_state.next_prod_id = max_id + 1
    conn.close()
    return True

def delete_product_from_db(p_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM products WHERE p_id = ?', (p_id,))
    c.execute('DELETE FROM variations WHERE p_id = ?', (p_id,))
    conn.commit()
    conn.close()

def clear_entire_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM products')
    c.execute('DELETE FROM variations')
    conn.commit()
    conn.close()

init_db()

# --- EXCHANGE RATES API ---
@st.cache_data(ttl=3600)
def fetch_base_rates():
    default_rates = {"THB": 4.73, "PHP": 2.65}
    rates_out = {}
    try:
        url_thb = "https://open.er-api.com/v6/latest/THB"
        req = urllib.request.Request(url_thb, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode())
            thb_jpy = data.get("rates", {}).get("JPY")
            if thb_jpy: rates_out["THB"] = round(thb_jpy, 4)
    except Exception:
        rates_out["THB"] = default_rates["THB"]

    try:
        url_php = "https://open.er-api.com/v6/latest/PHP"
        req = urllib.request.Request(url_php, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode())
            php_jpy = data.get("rates", {}).get("JPY")
            if php_jpy: rates_out["PHP"] = round(php_jpy, 4)
    except Exception:
        rates_out["PHP"] = default_rates["PHP"]

    return rates_out

SLS_RATES_THB = [
    (100, 93), (200, 105), (300, 129), (400, 153), (500, 177),
    (600, 200), (700, 230), (800, 250), (900, 270), (1000, 300),
    (1100, 321), (1500, 388), (2000, 508), (2100, 550), (2200, 585),
    (2300, 595), (2400, 615), (2500, 628), (3000, 748), (3500, 868),
    (4000, 988), (4500, 1108), (5000, 1228), (5500, 1348), (6000, 1468),
    (6500, 1588), (7000, 1708), (7500, 1828), (8000, 1948), (8500, 2068),
    (9000, 2188), (9500, 2308), (10000, 2428), (10500, 2548), (11000, 2668),
    (11500, 2788), (12000, 2908), (12500, 3028), (13000, 3148), (13500, 3268),
    (14000, 3388), (14500, 3508), (15000, 3628), (15500, 3748), (16000, 3868),
    (16500, 3988), (17000, 4108), (17500, 4228), (18000, 4348), (18500, 4468),
    (19000, 4588), (19500, 4708), (20000, 4828)
]

SLS_RATES_PHP = [
    (50, 76), (100, 101), (150, 126), (200, 151), (250, 176),
    (300, 201), (350, 226), (400, 251), (450, 276), (500, 301),
    (550, 326), (600, 351), (650, 376), (700, 401), (750, 426),
    (800, 451), (850, 476), (900, 501), (950, 526), (1000, 551),
    (1050, 576), (1100, 601), (1150, 626), (1200, 651), (1250, 676),
    (1300, 701), (1350, 726), (1400, 751), (1450, 776), (1500, 801),
    (1550, 826)
]

def get_sls_shipping_fee_thb(weight_g):
    for max_weight, fee in SLS_RATES_THB:
        if weight_g <= max_weight: return fee
    last_weight, last_fee = SLS_RATES_THB[-1]
    extra_steps = ((weight_g - last_weight) + 499) // 500
    return last_fee + (extra_steps * 120)

def get_sls_shipping_fee_php(weight_g):
    for max_weight, fee in SLS_RATES_PHP:
        if weight_g <= max_weight: return fee
    last_weight, last_fee = SLS_RATES_PHP[-1]
    extra_steps = ((weight_g - last_weight) + 49) // 50
    return last_fee + (extra_steps * 25)

def calculate_net_price(buying_price_jpy, weight_g, profit_rate_pct=30.0, currency="THB", rate_to_jpy=None):
    try:
        buying_price_jpy = float(buying_price_jpy) if buying_price_jpy is not None else 0.0
        weight_g = float(weight_g) if weight_g is not None else 0.0
        profit_rate_pct = float(profit_rate_pct) if profit_rate_pct is not None else 0.0
    except (ValueError, TypeError):
        return 0

    if buying_price_jpy <= 0 or weight_g <= 0: return 0

    margin_factor = 1.0 - (profit_rate_pct / 100.0)
    if margin_factor <= 0: margin_factor = 0.01

    if currency == "THB":
        rate = rate_to_jpy if rate_to_jpy else 4.73
        buying_price_thb = buying_price_jpy / rate if rate > 0 else 0
        sls_fee = get_sls_shipping_fee_thb(weight_g)
        transportation_jp_thb = 70.0
        net_price = (sls_fee + buying_price_thb + transportation_jp_thb) / margin_factor
        return round(net_price)

    elif currency == "PHP":
        rate = rate_to_jpy if rate_to_jpy else 2.65
        buying_price_php = buying_price_jpy / rate if rate > 0 else 0
        transportation_jp_php = 116.0
        sls_fee = get_sls_shipping_fee_php(weight_g)
        base_price = (sls_fee + buying_price_php + transportation_jp_php) / margin_factor
        cif_check = (base_price * 1.01) + (1369.0 * (weight_g / 1000.0))
        if cif_check >= 10000:
            custom_tax_part = 1369.0 * (weight_g / 1000.0) * 0.12
            net_price = (sls_fee + custom_tax_part + buying_price_php + transportation_jp_php) / 0.5788
        else:
            net_price = base_price
        return round(net_price)

    return 0

def on_field_change(p_id):
    save_product_to_db(p_id)

def create_blank_product(p_id):
    st.session_state[f"cat_{p_id}"] = ""
    st.session_state[f"psku_{p_id}"] = ""
    st.session_state[f"integ_{p_id}"] = ""
    st.session_state[f"brand_{p_id}"] = ""
    st.session_state[f"name_{p_id}"] = ""
    st.session_state[f"w_{p_id}"] = 0.0
    st.session_state[f"desc_{p_id}"] = ""
    st.session_state[f"cimg_{p_id}"] = ""
    st.session_state[f"v1n_{p_id}"] = ""
    st.session_state[f"v1o_{p_id}"] = ""
    st.session_state[f"v1i_{p_id}"] = ""
    st.session_state[f"v2n_{p_id}"] = ""
    st.session_state[f"v2o_{p_id}"] = ""
    st.session_state[f"df_data_{p_id}"] = pd.DataFrame()

def create_demo_product(p_id):
    st.session_state[f"cat_{p_id}"] = "120039"
    st.session_state[f"psku_{p_id}"] = "361086-18"
    st.session_state[f"integ_{p_id}"] = ""
    st.session_state[f"brand_{p_id}"] = "No Brand"
    st.session_state[f"name_{p_id}"] = T["default_pname"]
    st.session_state[f"w_{p_id}"] = 300.0
    st.session_state[f"desc_{p_id}"] = T["default_pdesc"]
    st.session_state[f"cimg_{p_id}"] = "https://example.com/cover.jpg, https://example.com/img1.jpg"
    st.session_state[f"v1n_{p_id}"] = T["default_v1_name"]
    st.session_state[f"v1o_{p_id}"] = T["default_v1_opts"]
    st.session_state[f"v1i_{p_id}"] = "https://example.com/wine.jpg, https://example.com/white.jpg"
    st.session_state[f"v2n_{p_id}"] = T["default_v2_name"]
    st.session_state[f"v2o_{p_id}"] = T["default_v2_opts"]

if "loaded_from_db" not in st.session_state:
    has_db_data = load_all_from_db()
    st.session_state.loaded_from_db = True
    if not has_db_data:
        st.session_state.products_list = [0]
        st.session_state.next_prod_id = 1
        create_demo_product(0)
        save_product_to_db(0)

# --- CONTROL PANEL ---
realtime_rates = fetch_base_rates()
st.subheader(T["calc_setting"])
col_cur, col_rate = st.columns(2)

with col_cur:
    currency = st.selectbox(T["currency_select"], ["THB", "PHP"], key="global_currency")

with col_rate:
    current_realtime_rate = realtime_rates.get(currency, 4.73 if currency == "THB" else 2.65)
    rate_jpy = st.number_input(
        T["rate_label"].format(curr=currency), 
        value=current_realtime_rate, 
        format="%.4f",
        help=T["rate_info"],
        key="global_rate"
    )
    st.caption(f"{T['rate_info']}: **1 {currency} = {rate_jpy} JPY**")

st.markdown("---")

col_btn1, col_btn2, col_space = st.columns([2, 2, 6])
with col_btn1:
    if st.button(T["add_product"]):
        new_id = st.session_state.next_prod_id
        st.session_state.next_prod_id += 1
        st.session_state.products_list.append(new_id)
        create_blank_product(new_id)
        save_product_to_db(new_id)
        st.rerun()

with col_btn2:
    if st.button(T["clear_all"], type="secondary"):
        clear_entire_db()
        for k in list(st.session_state.keys()):
            if k not in ["global_currency", "global_rate", "app_language"]:
                del st.session_state[k]
        st.session_state.products_list = [0]
        st.session_state.next_prod_id = 1
        create_demo_product(0)
        save_product_to_db(0)
        st.rerun()

updated_products_data = []
prod_id_to_remove = None

for idx, p_id in enumerate(st.session_state.products_list):
    st.markdown("---")
    col_title, col_del = st.columns([8, 2])
    
    title_text = st.session_state.get(f'name_{p_id}', '')
    display_title = title_text if title_text else f"{T['product_num']} {idx + 1}"

    with col_title:
        st.subheader(f"{T['product_num']} {idx + 1}: {display_title}")
    with col_del:
        if len(st.session_state.products_list) > 1:
            if st.button(f"{T['del_product']}", key=f"del_btn_{p_id}"):
                prod_id_to_remove = p_id

    c1, c2, c3 = st.columns(3)
    with c1:
        cat_id_val = st.text_input(T["cat_id"], key=f"cat_{p_id}", on_change=on_field_change, args=(p_id,))
        col_psku, col_integ = st.columns(2)
        with col_psku: psku_val = st.text_input(T["parent_sku"], key=f"psku_{p_id}", on_change=on_field_change, args=(p_id,))
        with col_integ: integ_val = st.text_input(T["integration_no"], key=f"integ_{p_id}", on_change=on_field_change, args=(p_id,))
        brand_val = st.text_input(T["brand"], key=f"brand_{p_id}", on_change=on_field_change, args=(p_id,))
    with c2:
        pname_val = st.text_input(T["p_name"], key=f"name_{p_id}", on_change=on_field_change, args=(p_id,))
        weight_val = st.number_input(T["weight"], step=10.0, format="%.1f", key=f"w_{p_id}", on_change=on_field_change, args=(p_id,))
    with c3:
        pdesc_val = st.text_area(T["p_desc"], key=f"desc_{p_id}", on_change=on_field_change, args=(p_id,))

    cimg_val = st.text_input(T["cover_img"], help=T["cover_img_help"], key=f"cimg_{p_id}", on_change=on_field_change, args=(p_id,))

    cv1, cv2 = st.columns(2)
    with cv1:
        v1n_val = st.text_input(T["v1_name"], key=f"v1n_{p_id}", on_change=on_field_change, args=(p_id,))
        v1o_val = st.text_input(T["v1_opts"], key=f"v1o_{p_id}", on_change=on_field_change, args=(p_id,))
        v1i_val = st.text_input(T["v1_imgs"], help=T["v1_imgs_help"], key=f"v1i_{p_id}", on_change=on_field_change, args=(p_id,))
    with cv2:
        v2n_val = st.text_input(T["v2_name"], key=f"v2n_{p_id}", on_change=on_field_change, args=(p_id,))
        v2o_val = st.text_input(T["v2_opts"], key=f"v2o_{p_id}", on_change=on_field_change, args=(p_id,))

    list_v1 = [x.strip() for x in v1o_val.split(",") if x.strip()]
    list_v2 = [x.strip() for x in v2o_val.split(",") if x.strip()] if v2n_val else [""]
    if not list_v2: list_v2 = [""]

    variations = list(itertools.product(list_v1, list_v2)) if list_v1 else []

    st.write(T["batch_title"])
    b_col1, b_col2, b_col3, b_col4 = st.columns(4)
    
    with b_col1:
        batch_cost = st.number_input(T["cost_col"], value=0, step=100, key=f"b_cost_{p_id}")
        apply_cost = st.button(f"{T['btn_apply']} {T['cost_col']}", key=f"btn_apply_cost_{p_id}")
    with b_col2:
        batch_weight = st.number_input(T["weight_col"], value=float(weight_val if weight_val else 0.0), step=10.0, format="%.1f", key=f"b_weight_{p_id}")
        apply_weight = st.button(f"{T['btn_apply']} {T['weight_col']}", key=f"btn_apply_weight_{p_id}")
    with b_col3:
        batch_profit = st.number_input(T["profit_col"], value=30.0, step=1.0, format="%.1f", key=f"b_profit_{p_id}")
        apply_profit = st.button(f"{T['btn_apply']} {T['profit_col']}", key=f"btn_apply_profit_{p_id}")
    with b_col4:
        batch_stock = st.number_input(T["stock_col"], value=2, step=1, key=f"b_stock_{p_id}")
        apply_stock = st.button(f"{T['btn_apply']} {T['stock_col']}", key=f"btn_apply_stock_{p_id}")

    df_key = f"df_data_{p_id}"
    existing_df = st.session_state.get(df_key, pd.DataFrame())

    existing_map = {}
    if isinstance(existing_df, pd.DataFrame) and not existing_df.empty:
        for _, r in existing_df.iterrows():
            existing_map[(str(r.get("Opt1", "")), str(r.get("Opt2", "")))] = r.to_dict()

    if variations:
        new_rows = []
        for opt1, opt2 in variations:
            key_tuple = (str(opt1), str(opt2))
            var_title = f"{opt1}" + (f" / {opt2}" if opt2 else "")
            sku_suffix = f"-{opt1}" + (f"-{opt2}" if opt2 else "")
            default_sku = f"{psku_val}{sku_suffix}" if psku_val else f"SKU{sku_suffix}"

            if key_tuple in existing_map:
                old_row = existing_map[key_tuple]
                old_row["Variation"] = var_title
                new_rows.append(old_row)
            else:
                new_rows.append({
                    "Variation": var_title, "SKU": default_sku,
                    "cost_jpy": int(batch_cost), "weight_g": float(batch_weight),
                    "profit_rate": float(batch_profit), "selling_price": 0,
                    "stock": int(batch_stock), "Opt1": str(opt1), "Opt2": str(opt2)
                })
        current_df = pd.DataFrame(new_rows)
    else:
        current_df = existing_df if isinstance(existing_df, pd.DataFrame) else pd.DataFrame()

    if not current_df.empty:
        if apply_cost:
            current_df["cost_jpy"] = int(batch_cost)
            st.session_state[df_key] = current_df
            save_variations_to_db(p_id, current_df)
            st.rerun()
        if apply_weight:
            current_df["weight_g"] = float(batch_weight)
            st.session_state[df_key] = current_df
            save_variations_to_db(p_id, current_df)
            st.rerun()
        if apply_profit:
            current_df["profit_rate"] = float(batch_profit)
            st.session_state[df_key] = current_df
            save_variations_to_db(p_id, current_df)
            st.rerun()
        if apply_stock:
            current_df["stock"] = int(batch_stock)
            st.session_state[df_key] = current_df
            save_variations_to_db(p_id, current_df)
            st.rerun()

    st.session_state[df_key] = current_df

    df_display = current_df.copy()
    if not df_display.empty:
        df_display["selling_price"] = df_display.apply(
            lambda row: calculate_net_price(
                buying_price_jpy=row.get("cost_jpy", 0),
                weight_g=row.get("weight_g", 0),
                profit_rate_pct=row.get("profit_rate", 0),
                currency=currency,
                rate_to_jpy=rate_jpy
            ), axis=1
        )

    st.write(T["grid_title"])
    edited_df = st.data_editor(
        df_display,
        column_config={
            "Variation": st.column_config.Column(disabled=True),
            "SKU": st.column_config.TextColumn(T["sku_col"], disabled=False),
            "cost_jpy": st.column_config.NumberColumn(T["cost_col"], min_value=0, format="%d ¥"),
            "weight_g": st.column_config.NumberColumn(T["weight_col"], min_value=0.0, format="%.1f g"),
            "profit_rate": st.column_config.NumberColumn(T["profit_col"], min_value=0.0, max_value=99.0, format="%.1f %%"),
            "selling_price": st.column_config.NumberColumn(f"{T['price_col']} ({currency})", disabled=True, format="%d " + currency),
            "stock": st.column_config.NumberColumn(T["stock_col"], min_value=0, format="%d"),
            "Opt1": None, "Opt2": None
        },
        hide_index=True,
        key=f"editor_{p_id}"
    )

    if not edited_df.empty:
        for col in ["SKU", "cost_jpy", "weight_g", "profit_rate", "stock"]:
            if col in edited_df.columns:
                st.session_state[df_key][col] = edited_df[col]
        save_variations_to_db(p_id, st.session_state[df_key])

    v1_imgs_list = [x.strip() for x in v1i_val.split(",") if x.strip()]
    updated_products_data.append({
        "cat_id": cat_id_val, "p_sku": psku_val, "integration_no": integ_val,
        "brand": brand_val, "p_name": pname_val, "weight": weight_val if weight_val else 0,
        "p_desc": pdesc_val, "cover_img": cimg_val, "v1_name": v1n_val,
        "v1_opts_list": list_v1, "v1_imgs": v1_imgs_list, "v2_name": v2n_val,
        "variations_table": st.session_state[df_key], "cost_key": "cost_jpy", "weight_key": "weight_g",
        "price_key": "selling_price", "stock_key": "stock"
    })

if prod_id_to_remove is not None:
    st.session_state.products_list.remove(prod_id_to_remove)
    if f"df_data_{prod_id_to_remove}" in st.session_state:
        del st.session_state[f"df_data_{prod_id_to_remove}"]
    delete_product_from_db(prod_id_to_remove)
    st.rerun()

# --- EXPORT EXCEL ---
st.markdown("---")
if st.button(T["btn_generate"], type="primary", use_container_width=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    header_row1 = [
        "Category", "Product Name", "Product Description", "Maximum Purchase Quantity", 
        "Maximum Purchase Quantity - Start Date", "Maximum Purchase Quantity - Time Period (in Days)", 
        "Maximum Purchase Quantity - End Date", "Minimum Purchase Quantity", "Parent SKU", 
        "Variation Integration No.", "Variation Name1", "Option for Variation 1", "Image per Variation", 
        "Variation Name2", "Option for Variation 2", "Price", "Stock", "SKU", "Size Chart Template", 
        "Size Chart Image", "Cover image", "Item Image 1", "Item Image 2", "Item Image 3", 
        "Item Image 4", "Item Image 5", "Item Image 6", "Item Image 7", "Item Image 8", "Weight", 
        "Length", "Width", "Height", "International Express - ส่งจากต่างประเทศ (Japan)", 
        "Pre-order DTS", "Fail Reason"
    ]
    header_row2 = ["Optional", "Mandatory", "Mandatory", "Optional", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Optional", "Optional", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Mandatory", "Conditional Mandatory", "Optional", "Conditional Mandatory", "Conditional Mandatory", "Mandatory", "Optional", "Optional", "Optional", "Optional", "Optional", "Optional", "Optional", "Optional", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Conditional Mandatory", "Optional", "Conditional Mandatory", "Optional"]
    header_row3 = ["Indicate the appropriate category ID for each product.", "Product name should include product brand and model.", "A good product description enhances the quality of your listing.", "[Per Order + Per Time Period]", "Please select a MaxPQ start date.", "[Per Time Period only]", "Please select a MaxPQ end date", "MPQ is an item level field.", "Parent SKU is used to identify parent products.", "Mandatory for products with variations.", "Please indicate the first variation name.", "Indicate the first variation value.", "Upload an image per variation.", "Indicate the second variation name.", "Please indicate the second variation value.", "Input your product price.", "Input your product stock.", "SKU is a unique identifier.", "Please enter the size chart template ID.", "You only need to fill in either size chart.", "Upload the URL of your main product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Enter the URL of this product image.", "Input your product weight.", "Fill up all dimensions.", "Input your product width.", "Input your product height.", "Please toggle 'on'", "Pre-order DTS range", ""]
    header_row4 = ["Choose your desired category ID from the Category Tree.", "Please input 20 to 255 characters.", "Please input 60 to 5000 characters.", "Please input from 1 to 999,999.", "YYYY-MM-DD", "Please input from 1 to 365.", "YYYY-MM-DD", "Minimum purchase quantity only can be a positive integer.", "Please input 1-100 characters.", "Input 1 to 100 characters.", "Input from 1 to 14 characters.", "Input from 1 to 30 characters.", "Enter the URL of this product image.", "Input from 1 to 14 characters.", "Input from 1 to 30 characters.", "Input price.", "Input stock.", "Input less than 100 characters.", "Please enter size chart template ID.", "Size: Max 2Mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Size: max 2.0mb", "Please input 0.00 to 1000000.00", "Please input 0 to 1000000", "Please input 0 to 1000000", "Please input 0 to 1000000", "On/Off", "Pre-order DTS range", ""]
    header_row5 = ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Format accepted: PDF, JPG, JPEG, PNG", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "Format accepted: JPG, JPEG, PNG.", "", "", "", "", "", "", ""]

    ws.append(header_row1)
    ws.append(header_row2)
    ws.append(header_row3)
    ws.append(header_row4)
    ws.append(header_row5)

    start_row = 6

    for p_data in updated_products_data:
        cat_id = p_data["cat_id"]
        p_sku = p_data["p_sku"]
        integration_no = p_data["integration_no"]
        p_name = p_data["p_name"]
        p_desc = p_data["p_desc"]
        cover_img = p_data["cover_img"]
        v1_name = p_data["v1_name"]
        v2_name = p_data["v2_name"]
        v1_opts_list = p_data["v1_opts_list"]
        v1_imgs = p_data["v1_imgs"]
        df_vars = p_data["variations_table"]
        
        pr_k = p_data["price_key"]
        st_k = p_data["stock_key"]
        w_k = p_data["weight_key"]

        v1_img_dict = {}
        for i, opt1_val in enumerate(v1_opts_list):
            v1_img_dict[opt1_val] = v1_imgs[i] if i < len(v1_imgs) else ""

        if df_vars is not None and not df_vars.empty:
            for idx, row in df_vars.iterrows():
                current_row = start_row
                
                calc_price = calculate_net_price(
                    buying_price_jpy=row.get("cost_jpy", 0),
                    weight_g=row.get("weight_g", 0),
                    profit_rate_pct=row.get("profit_rate", 0),
                    currency=currency,
                    rate_to_jpy=rate_jpy
                )

                ws.cell(row=current_row, column=1, value=cat_id)
                ws.cell(row=current_row, column=2, value=p_name if idx == 0 else "")
                ws.cell(row=current_row, column=3, value=p_desc if idx == 0 else "")
                ws.cell(row=current_row, column=9, value=p_sku)
                ws.cell(row=current_row, column=10, value=integration_no)
                ws.cell(row=current_row, column=11, value=v1_name)
                ws.cell(row=current_row, column=12, value=row.get("Opt1", ""))
                ws.cell(row=current_row, column=13, value=v1_img_dict.get(row.get("Opt1", ""), ""))
                
                if v2_name and row.get("Opt2", ""):
                    ws.cell(row=current_row, column=14, value=v2_name)
                    ws.cell(row=current_row, column=15, value=row.get("Opt2", ""))
                    
                ws.cell(row=current_row, column=16, value=calc_price)
                ws.cell(row=current_row, column=17, value=row.get(st_k, 0))
                ws.cell(row=current_row, column=18, value=row.get("SKU", ""))
                
                if idx == 0:
                    cover_imgs_list = [x.strip() for x in cover_img.split(",") if x.strip()]
                    for img_i, img_url in enumerate(cover_imgs_list):
                        if img_i < 9:
                            ws.cell(row=current_row, column=21 + img_i, value=img_url)

                    ws.cell(row=current_row, column=30, value=row.get(w_k, 0) / 1000.0)
                    ws.cell(row=current_row, column=34, value="On")

                start_row += 1

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="000000")
    data_font = Font(name="Calibri", size=10, color="000000")
    wrap_alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 20

    for r in range(6, start_row):
        ws.row_dimensions[r].height = 28

    max_col = len(header_row1)
    for r in range(1, start_row):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if r <= 5:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                cell.font = data_font
                cell.alignment = wrap_alignment

    column_widths = {
        "A": 15, "B": 35, "C": 45, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15,
        "I": 25, "J": 25, "K": 20, "L": 20, "M": 30, "N": 20, "O": 20, "P": 15,
        "Q": 12, "R": 25, "S": 20, "T": 30, "U": 35, "V": 30, "W": 30, "X": 30,
        "Y": 30, "Z": 30, "AA": 30, "AB": 30, "AC": 30, "AD": 15, "AE": 12,
        "AF": 12, "AG": 12, "AH": 20, "AI": 15, "AJ": 20
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    st.success(T["success_msg"].format(count=len(updated_products_data)))
    st.download_button(
        label=T["btn_download"],
        data=output,
        file_name=f"Shopee_Mass_Upload_{currency}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )